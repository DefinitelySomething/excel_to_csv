from openpyxl import load_workbook
from openpyxl import workbook
import csv
import sys
import os

excel = sys.argv[1]

drive = 'c:\\'
folder_parent = 'tools'
folder = 'fastload'

location = os.path.join(drive, folder_parent, folder)

filename = "LOAD.csv"

file = os.path.join(location,filename)

if not os.path.exists(location):
    os.makedirs(location)

if os.path.exists(file):
    os.remove(file)

wb = load_workbook(filename = excel)

first_sheet = wb[wb.sheetnames[0]]

first_sheet.delete_rows(0)

with open(file,'w', newline='') as csvfile:
    csvwriter = csv.writer(csvfile,
                           delimiter='|',
                           )
    for row in  first_sheet.rows:
        csvwriter.writerow([cell.value for cell in row])

print ('success!!')




##to compile
##pip install pyinstaller
##from command line of saved python file: pyinstaller -- onefile --console 
